from django.views.generic import ListView, CreateView, DetailView, UpdateView, DeleteView, View, TemplateView
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.urls import reverse_lazy
from django.db.models import Q
from django.http import HttpResponse, JsonResponse
from django.utils import timezone
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from .models import Vehicle, AuditLog, UserProfile


class VehicleListView(LoginRequiredMixin, ListView):
    """
    Список всех транспортных средств с поиском и пагинацией.

    Поиск работает по:
    - Гаражному номеру
    - VIN номеру
    - Государственному регистрационному номеру (ГРЗ)
    - Марке и модели
    - Номеру полиса ОСАГО
    - Номеру диагностической карты
    (все поиски нечувствительны к регистру)

    Требует аутентификации.
    """
    model = Vehicle
    template_name = 'vehicle_list.html'
    context_object_name = 'vehicles'
    paginate_by = 10
    login_url = 'admin:login'

    def get_queryset(self):
        # Переопределяем queryset, чтобы добавить фильтрацию по поисковому запросу
        queryset = super().get_queryset().filter(is_archived=False)
        q = self.request.GET.get('q')
        if q:
            # Расширенный поиск по множеству полей
            queryset = queryset.filter(
                Q(garage_number__icontains=q) |           # Гаражный номер
                Q(vin__icontains=q) |                     # VIN номер
                Q(grz_number__icontains=q) |              # Номер ГРЗ (АА123БВ)
                Q(grz_series__icontains=q) |              # Серия ГРЗ (78)
                Q(brand_model__name__icontains=q) |       # Марка и модель
                Q(osago_number__icontains=q) |            # Номер ОСАГО
                Q(diagnostic_card_number__icontains=q) |  # Номер диагностической карты
                Q(reg_certificate_number__icontains=q)    # Номер свидетельства регистрации
            )
        return queryset


class VehicleCreateView(LoginRequiredMixin, UserPassesTestMixin, CreateView):
    """
    Форма создания нового транспортного средства.

    Только для пользователей в группах: Администратор, Специалист
    """
    model = Vehicle
    template_name = 'vehicle_form.html'
    fields = [
        # Основные реквизиты
        'garage_number', 'brand_model', 'vin', 'photo',
        # Характеристики кузова
        'color', 'vehicle_type', 'category', 'fuel_type', 'body_type',
        # Паспорт ТС
        'pts_type', 'pts_series', 'pts_number', 'pts_date', 'pts_scan',
        'epts_number', 'epts_date', 'epts_scan',
        # Регистрационные данные ГИБДД
        'grz_series', 'grz_number', 'grz_date',
        'reg_certificate_series', 'reg_certificate_number', 'reg_certificate_date', 'reg_certificate_scan',
        'osago_number', 'osago_date_issued', 'osago_date_expiry', 'osago_cost', 'osago_scan',
        'diagnostic_card_number', 'diagnostic_card_date', 'diagnostic_card_expiry', 'diagnostic_card_scan',
        # Техническое снятие с учета
        'dismission_date',
        # Характеристики двигателя
        'engine_volume', 'engine_power_hp', 'engine_power_kw', 'engine_model_number',
        # Номера конструктивных элементов
        'chassis_number', 'body_number',
        # Год выпуска и характеристики
        'year_of_manufacture', 'axles_count',
        # Масса и грузоподъемность
        'mass_kg', 'empty_mass', 'cargo_capacity', 'passenger_capacity',
        # Габариты и емкости
        'bus_length', 'bucket_capacity', 'fuel_tank_capacity',
        # Топливо
        'fuel_grade', 'fuel_consumption_type',
        'fuel_consumption_summer_city', 'fuel_consumption_summer_highway',
        'fuel_consumption_winter_city', 'fuel_consumption_winter_highway',
        'fuel_consumption_per_hour',
        # Классификация для разных ведомств
        'vehicle_type_for_military', 'vehicle_type_for_tax', 'initial_cost',
        # Общие данные и статус
        'inventory_number', 'motorpool', 'direction', 'ts_status',
        # Пробег и моточасы
        'mileage_km', 'mileage_type', 'motor_hours', 'to2_periodicity',
        # ГБО (газобаллонное оборудование)
        'gbo_present', 'gbo_type', 'gbo_volume', 'gbo_installation_date',
        'gbo_cylinders_count', 'gbo_total_capacity',
        'gbo_inspection_date', 'gbo_next_inspection_date', 'gbo_inspection_scan',
        # ССМТ и другое оборудование
        'has_ssmt', 'has_tachograph',
        'tachograph_calibration_date', 'tachograph_next_calibration_date',
        'has_lifting_equipment',
        # Техническое обслуживание
        'dprg_inspection_date', 'dprg_next_inspection_date', 'dprg_inspection_scan',
        # Бухгалтерские данные
        'okof_code', 'date_in_service', 'depreciation_rate',
        'accumulated_depreciation', 'residual_value',
        # Старые поля сроков (для совместимости)
        'to2_period_days', 'to3_period_days', 'tech_inspection_expiry', 'insurance_expiry'
    ]
    login_url = 'admin:login'

    def test_func(self):
        """Проверка: может ли пользователь редактировать"""
        groups = [g.name for g in self.request.user.groups.all()]
        return 'Администратор' in groups or 'Специалист' in groups or 'Руководитель' in groups


class VehicleDetailView(LoginRequiredMixin, DetailView):
    """
    Подробный просмотр одного транспортного средства.

    Отображает все реквизиты карточки со ссылками на редактирование и удаление
    (в зависимости от прав пользователя).
    """
    model = Vehicle
    template_name = 'vehicle_detail.html'
    context_object_name = 'vehicle'
    login_url = 'admin:login'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Добавляем информацию о правах в контекст
        groups = [g.name for g in self.request.user.groups.all()]
        context['can_edit'] = 'Администратор' in groups or 'Специалист' in groups or 'Руководитель' in groups
        context['can_delete'] = 'Администратор' in groups or 'Руководитель' in groups
        return context


class VehicleUpdateView(LoginRequiredMixin, UserPassesTestMixin, UpdateView):
    """
    Форма редактирования существующего транспортного средства.

    Только для пользователей в группах: Администратор, Специалист, Руководитель
    """
    model = Vehicle
    template_name = 'vehicle_form.html'
    fields = [
        # Основные реквизиты
        'garage_number', 'brand_model', 'vin', 'photo',
        # Характеристики кузова
        'color', 'vehicle_type', 'category', 'fuel_type', 'body_type',
        # Паспорт ТС
        'pts_type', 'pts_series', 'pts_number', 'pts_date', 'pts_scan',
        'epts_number', 'epts_date', 'epts_scan',
        # Регистрационные данные ГИБДД
        'grz_series', 'grz_number', 'grz_date',
        'reg_certificate_series', 'reg_certificate_number', 'reg_certificate_date', 'reg_certificate_scan',
        'osago_number', 'osago_date_issued', 'osago_date_expiry', 'osago_cost', 'osago_scan',
        'diagnostic_card_number', 'diagnostic_card_date', 'diagnostic_card_expiry', 'diagnostic_card_scan',
        # Техническое снятие с учета
        'dismission_date',
        # Характеристики двигателя
        'engine_volume', 'engine_power_hp', 'engine_power_kw', 'engine_model_number',
        # Номера конструктивных элементов
        'chassis_number', 'body_number',
        # Год выпуска и характеристики
        'year_of_manufacture', 'axles_count',
        # Масса и грузоподъемность
        'mass_kg', 'empty_mass', 'cargo_capacity', 'passenger_capacity',
        # Габариты и емкости
        'bus_length', 'bucket_capacity', 'fuel_tank_capacity',
        # Топливо
        'fuel_grade', 'fuel_consumption_type',
        'fuel_consumption_summer_city', 'fuel_consumption_summer_highway',
        'fuel_consumption_winter_city', 'fuel_consumption_winter_highway',
        'fuel_consumption_per_hour',
        # Классификация для разных ведомств
        'vehicle_type_for_military', 'vehicle_type_for_tax', 'initial_cost',
        # Общие данные и статус
        'inventory_number', 'motorpool', 'direction', 'ts_status',
        # Пробег и моточасы
        'mileage_km', 'mileage_type', 'motor_hours', 'to2_periodicity',
        # ГБО (газобаллонное оборудование)
        'gbo_present', 'gbo_type', 'gbo_volume', 'gbo_installation_date',
        'gbo_cylinders_count', 'gbo_total_capacity',
        'gbo_inspection_date', 'gbo_next_inspection_date', 'gbo_inspection_scan',
        # ССМТ и другое оборудование
        'has_ssmt', 'has_tachograph',
        'tachograph_calibration_date', 'tachograph_next_calibration_date',
        'has_lifting_equipment',
        # Техническое обслуживание
        'dprg_inspection_date', 'dprg_next_inspection_date', 'dprg_inspection_scan',
        # Бухгалтерские данные
        'okof_code', 'date_in_service', 'depreciation_rate',
        'accumulated_depreciation', 'residual_value',
        # Старые поля сроков (для совместимости)
        'to2_period_days', 'to3_period_days', 'tech_inspection_expiry', 'insurance_expiry'
    ]
    success_url = reverse_lazy('cards:vehicle_list')
    login_url = 'admin:login'

    def test_func(self):
        """Проверка: может ли пользователь редактировать"""
        groups = [g.name for g in self.request.user.groups.all()]
        return 'Администратор' in groups or 'Специалист' in groups or 'Руководитель' in groups


class VehicleDeleteView(LoginRequiredMixin, UserPassesTestMixin, DeleteView):
    """
    Удаление транспортного средства с подтверждением.

    Только для пользователей в группах: Администратор, Руководитель
    """
    model = Vehicle
    template_name = 'vehicle_confirm_delete.html'
    success_url = reverse_lazy('cards:vehicle_list')
    login_url = 'admin:login'

    def test_func(self):
        """Проверка: может ли пользователь удалять"""
        groups = [g.name for g in self.request.user.groups.all()]
        return 'Администратор' in groups or 'Руководитель' in groups


class VehicleExportView(LoginRequiredMixin, View):
    """
    Экспорт списка транспортных средств в Excel.

    Поддерживает фильтрацию по поисковому запросу 'q' (аналогично VehicleListView).
    Требует аутентификации.
    """
    login_url = 'admin:login'

    def get_queryset(self):
        """Получить отфильтрованный список ТС (с учётом поиска)"""
        queryset = Vehicle.objects.all()
        q = self.request.GET.get('q')
        if q:
            queryset = queryset.filter(
                Q(garage_number__icontains=q) |
                Q(vin__icontains=q) |
                Q(grz_number__icontains=q) |
                Q(grz_series__icontains=q) |
                Q(brand_model__name__icontains=q) |
                Q(osago_number__icontains=q) |
                Q(diagnostic_card_number__icontains=q) |
                Q(reg_certificate_number__icontains=q)
            )
        return queryset.order_by('garage_number')

    def get(self, request, *args, **kwargs):
        """Генерирует Excel файл и возвращает его как ответ"""
        vehicles = self.get_queryset()

        # Создаём новую рабочую книгу
        wb = Workbook()
        ws = wb.active
        ws.title = "Транспортные средства"

        # Стили
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Заголовки колонок
        headers = [
            "Гаражный номер",
            "Марка и модель",
            "VIN",
            "Цвет",
            "Тип кузова",
            "Тип топлива",
            "Категория",
            "Масса (кг)",
            "Пробег (км)",
            "Объем двигателя (л)",
            "Мощность (л.с.)",
            "Осей",
            "ПТС/ЭПТС",
            "ГРЗ",
            "ОСАГО (номер)",
            "ОСАГО (до)",
            "Диагностическая карта (номер)",
            "Диагностическая карта (до)",
            "Техосмотр (до)",
            "Статус документов"
        ]

        # Вписываем заголовки в первую строку
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border

        # Вписываем данные
        for row_idx, vehicle in enumerate(vehicles, 2):
            data = [
                vehicle.garage_number,
                str(vehicle.brand_model),
                vehicle.vin,
                str(vehicle.color) if vehicle.color else "",
                str(vehicle.body_type) if vehicle.body_type else "",
                str(vehicle.fuel_type) if vehicle.fuel_type else "",
                str(vehicle.category) if vehicle.category else "",
                vehicle.mass_kg or "",
                vehicle.mileage_km or "",
                vehicle.engine_volume or "",
                vehicle.engine_power_hp or "",
                vehicle.axles_count or "",
                vehicle.get_pts_type_display() if vehicle.pts_type else "",
                f"{vehicle.grz_series}{vehicle.grz_number}" if vehicle.grz_series and vehicle.grz_number else "",
                vehicle.osago_number or "",
                vehicle.osago_date_expiry.strftime("%d.%m.%Y") if vehicle.osago_date_expiry else "",
                vehicle.diagnostic_card_number or "",
                vehicle.diagnostic_card_expiry.strftime("%d.%m.%Y") if vehicle.diagnostic_card_expiry else "",
                vehicle.tech_inspection_expiry.strftime("%d.%m.%Y") if vehicle.tech_inspection_expiry else "",
                vehicle.get_expiry_status()[1]  # Текстовое описание статуса
            ]

            for col_idx, value in enumerate(data, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = value
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                cell.border = border

        # Автоматическая ширина колонок
        ws.column_dimensions['A'].width = 14
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 17
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 14
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 12
        ws.column_dimensions['I'].width = 12
        ws.column_dimensions['J'].width = 12
        ws.column_dimensions['K'].width = 12
        ws.column_dimensions['L'].width = 10
        ws.column_dimensions['M'].width = 12
        ws.column_dimensions['N'].width = 12
        ws.column_dimensions['O'].width = 14
        ws.column_dimensions['P'].width = 14
        ws.column_dimensions['Q'].width = 18
        ws.column_dimensions['R'].width = 18
        ws.column_dimensions['S'].width = 14
        ws.column_dimensions['T'].width = 20

        # Высота строки заголовка
        ws.row_dimensions[1].height = 28

        # Возвращаем файл как HTTP ответ
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        response['Content-Disposition'] = f'attachment; filename="vehicles_{timestamp}.xlsx"'
        wb.save(response)
        return response


class VehicleArchiveView(LoginRequiredMixin, UserPassesTestMixin, View):
    """
    Архивирование транспортного средства.

    Только для пользователей в группах: Администратор, Руководитель
    """
    login_url = 'admin:login'

    def test_func(self):
        """Проверка: может ли пользователь архивировать"""
        groups = [g.name for g in self.request.user.groups.all()]
        return 'Администратор' in groups or 'Руководитель' in groups

    def post(self, request, pk, *args, **kwargs):
        """Архивирует ТС и логирует действие"""
        try:
            vehicle = Vehicle.objects.get(pk=pk)
            vehicle.is_archived = True
            vehicle.archived_at = timezone.now()
            vehicle.archived_by = request.user
            vehicle.save()

            # Логируем архивирование
            AuditLog.objects.create(
                user_fio=request.user.profile.fio if hasattr(request.user, 'profile') and request.user.profile else request.user.username,
                username=request.user.username,
                action='UPDATE',
                model_name='Vehicle',
                object_id=vehicle.id,
                object_description=str(vehicle),
                changes='Архивировано (снято с учета)',
                ip_address=self._get_client_ip(request)
            )

            return JsonResponse({'status': 'success', 'message': 'ТС архивировано'})
        except Vehicle.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'ТС не найдено'}, status=404)

    def _get_client_ip(self, request):
        """Получить IP адрес клиента"""
        x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
        if x_forwarded_for:
            ip = x_forwarded_for.split(',')[0]
        else:
            ip = request.META.get('REMOTE_ADDR')
        return ip


class VehicleUnarchiveView(LoginRequiredMixin, UserPassesTestMixin, View):
    """
    Восстановление транспортного средства из архива.

    Только для пользователей в группах: Администратор, Руководитель
    """
    login_url = 'admin:login'

    def test_func(self):
        """Проверка: может ли пользователь восстанавливать из архива"""
        groups = [g.name for g in self.request.user.groups.all()]
        return 'Администратор' in groups or 'Руководитель' in groups

    def post(self, request, pk, *args, **kwargs):
        """Восстанавливает ТС из архива и логирует действие"""
        try:
            vehicle = Vehicle.objects.get(pk=pk)
            vehicle.is_archived = False
            vehicle.archived_at = None
            vehicle.archived_by = None
            vehicle.save()

            # Логируем восстановление
            AuditLog.objects.create(
                user_fio=request.user.profile.fio if hasattr(request.user, 'profile') and request.user.profile else request.user.username,
                username=request.user.username,
                action='UPDATE',
                model_name='Vehicle',
                object_id=vehicle.id,
                object_description=str(vehicle),
                changes='Восстановлено из архива (возвращено в учет)',
                ip_address=self._get_client_ip(request)
            )

            return JsonResponse({'status': 'success', 'message': 'ТС восстановлено'})
        except Vehicle.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'ТС не найдено'}, status=404)

    def _get_client_ip(self, request):
        """Получить IP адрес клиента"""
        x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
        if x_forwarded_for:
            ip = x_forwarded_for.split(',')[0]
        else:
            ip = request.META.get('REMOTE_ADDR')
        return ip


# ===== ДИРЕКТОРИЯ СОТРУДНИКОВ =====

class UserDirectoryView(LoginRequiredMixin, ListView):
    """
    Директория всех сотрудников с поиском и фильтрацией.

    Отображает:
    - ФИО
    - Табельный номер
    - Должность
    - Группа доступа
    - Дата создания

    Требует аутентификации.
    """
    model = UserProfile
    template_name = 'cards/user_directory.html'
    context_object_name = 'users'
    paginate_by = 20
    login_url = 'admin:login'

    def get_queryset(self):
        """Получить список сотрудников с фильтрацией"""
        queryset = UserProfile.objects.all()

        # Поиск по ФИО, табельному номеру или должности
        q = self.request.GET.get('q')
        if q:
            queryset = queryset.filter(
                Q(fio__icontains=q) |
                Q(tab_number__icontains=q) |
                Q(position__icontains=q) |
                Q(user__username__icontains=q)
            )

        return queryset.select_related('user').order_by('fio')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Добавляем информацию о группах для каждого сотрудника
        for user_profile in context['users']:
            user_profile.groups = user_profile.user.groups.all()

        return context


class EmployeeImportView(LoginRequiredMixin, UserPassesTestMixin, View):
    """
    Форма для импорта сотрудников из CSV файла.

    Только для администраторов.
    Поддерживаемый формат CSV:
    ФИО,Табельный номер,Должность,Логин,Группа

    Примеры:
    Иван Иванов,001,Администратор,iivanov,Администратор
    Мария Сидорова,002,Специалист,msidorova,Специалист
    """
    login_url = 'admin:login'
    template_name = 'cards/employee_import.html'

    def test_func(self):
        """Только администраторы"""
        groups = [g.name for g in self.request.user.groups.all()]
        return 'Администратор' in groups

    def get(self, request, *args, **kwargs):
        """Отобразить форму импорта"""
        from django.shortcuts import render
        return render(request, self.template_name)

    def post(self, request, *args, **kwargs):
        """Обработать загрузку CSV файла"""
        import csv
        from django.contrib.auth.models import User, Group
        from django.shortcuts import render

        if 'csv_file' not in request.FILES:
            return render(request, self.template_name, {'error': 'Файл не выбран'})

        csv_file = request.FILES['csv_file']

        if not csv_file.name.endswith('.csv'):
            return render(request, self.template_name, {'error': 'Файл должен быть в формате CSV'})

        try:
            # Декодируем файл
            stream = csv_file.read().decode('UTF8').splitlines()
            reader = csv.DictReader(stream)

            # Проверяем заголовки
            required_fields = {'ФИО', 'Табельный номер', 'Логин', 'Группа'}
            if not reader.fieldnames or not required_fields.issubset(set(reader.fieldnames)):
                return render(
                    request,
                    self.template_name,
                    {'error': f'CSV должен содержать столбцы: {", ".join(required_fields)}'}
                )

            created_count = 0
            updated_count = 0
            errors = []

            for row_num, row in enumerate(reader, start=2):
                try:
                    fio = row.get('ФИО', '').strip()
                    tab_number = row.get('Табельный номер', '').strip()
                    position = row.get('Должность', 'Пользователь').strip()
                    username = row.get('Логин', '').strip()
                    group_name = row.get('Группа', 'Пользователь').strip()

                    if not fio or not tab_number or not username:
                        errors.append(f'Строка {row_num}: отсутствуют обязательные поля')
                        continue

                    # Получаем или создаем группу
                    group, _ = Group.objects.get_or_create(name=group_name)

                    # Получаем или создаем пользователя
                    user, user_created = User.objects.get_or_create(username=username)
                    user.first_name = fio.split()[0] if fio else ''
                    user.last_name = ' '.join(fio.split()[1:]) if len(fio.split()) > 1 else ''
                    user.save()

                    # Добавляем в группу если еще не добавлен
                    if not user.groups.filter(pk=group.pk).exists():
                        user.groups.add(group)

                    # Получаем или создаем профиль
                    profile, profile_created = UserProfile.objects.get_or_create(
                        user=user,
                        defaults={
                            'fio': fio,
                            'tab_number': tab_number,
                            'position': position,
                        }
                    )

                    # Обновляем существующий профиль
                    if not profile_created:
                        profile.fio = fio
                        profile.tab_number = tab_number
                        profile.position = position
                        profile.save()
                        updated_count += 1
                    else:
                        created_count += 1

                except Exception as e:
                    errors.append(f'Строка {row_num}: {str(e)}')

            message = f'Успешно: создано {created_count} сотр., обновлено {updated_count} сотр.'
            if errors:
                message += f' Ошибок: {len(errors)}'
                return render(request, self.template_name, {
                    'success': message,
                    'errors': errors
                })

            return render(request, self.template_name, {'success': message})

        except Exception as e:
            return render(request, self.template_name, {
                'error': f'Ошибка при обработке файла: {str(e)}'
            })

# ===== ТЕХНИЧЕСКОЕ ОБСЛУЖИВАНИЕ (ТО-2) =====

class MaintenanceScheduleView(LoginRequiredMixin, ListView):
    """
    Список расписаний техническое обслуживание ТО-2.

    Показывает:
    - Транспортного средства, нуждающиеся в ТО-2
    - Статус выполнения
    - Дату следующей ТО-2
    - Фильтрация по статусу

    Требует аутентификации.
    """
    from .models import MaintenanceSchedule

    model = MaintenanceSchedule
    template_name = 'cards/maintenance_schedule.html'
    context_object_name = 'schedules'
    paginate_by = 20
    login_url = 'admin:login'

    def get_queryset(self):
        """Получить расписание с фильтрацией"""
        queryset = MaintenanceSchedule.objects.select_related('vehicle').all()

        # Фильтрация по статусу
        status = self.request.GET.get('status')
        if status:
            queryset = queryset.filter(status=status)

        # Фильтрация по ТС (поиск)
        q = self.request.GET.get('q')
        if q:
            queryset = queryset.filter(
                Q(vehicle__garage_number__icontains=q) |
                Q(vehicle__brand_model__name__icontains=q)
            )

        return queryset.order_by('next_maintenance_date')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Добавляем статистику
        from .models import MaintenanceSchedule as MS

        context['total_count'] = MS.objects.count()
        context['scheduled_count'] = MS.objects.filter(status='scheduled').count()
        context['in_progress_count'] = MS.objects.filter(status='in_progress').count()
        context['overdue_count'] = MS.objects.filter(status='overdue').count()
        context['completed_count'] = MS.objects.filter(status='completed').count()

        return context


# ===== СИСТЕМА (НАСТРОЙКИ И ИНФОРМАЦИЯ) =====

class SettingsView(LoginRequiredMixin, TemplateView):
    """
    Страница настроек приложения.

    Позволяет пользователям настраивать:
    - Автозапуск приложения
    - Оповещения в WC03P
    - Оповещения в ПП
    - Выбор шрифта (Times New Roman/Arial)
    - Размер шрифта (6-72)

    Требует аутентификации.
    """
    template_name = 'cards/settings.html'
    login_url = 'admin:login'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Получаем текущие настройки из сессии
        context['autostart'] = self.request.session.get('autostart', True)
        context['notify_wc03p'] = self.request.session.get('notify_wc03p', False)
        context['notify_app'] = self.request.session.get('notify_app', True)
        context['font_family'] = self.request.session.get('font_family', 'Arial')
        context['font_size'] = self.request.session.get('font_size', 14)

        return context

    def post(self, request, *args, **kwargs):
        """Сохранить настройки в сессию"""
        from django.shortcuts import redirect

        # Сохраняем настройки
        request.session['autostart'] = request.POST.get('autostart') == 'on'
        request.session['notify_wc03p'] = request.POST.get('notify_wc03p') == 'on'
        request.session['notify_app'] = request.POST.get('notify_app') == 'on'
        request.session['font_family'] = request.POST.get('font_family', 'Arial')
        request.session['font_size'] = int(request.POST.get('font_size', 14))

        # Логируем действие
        from .models import AuditLog
        AuditLog.objects.create(
            user_fio=request.user.profile.fio if hasattr(request.user, 'profile') else request.user.username,
            username=request.user.username,
            action='UPDATE',
            model_name='Settings',
            object_description='Пользовательские настройки приложения',
            ip_address=self._get_client_ip(request)
        )

        return redirect('cards:settings')

    @staticmethod
    def _get_client_ip(request):
        """Получить IP адрес клиента"""
        x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
        if x_forwarded_for:
            ip = x_forwarded_for.split(',')[0]
        else:
            ip = request.META.get('REMOTE_ADDR')
        return ip


class AboutView(LoginRequiredMixin, TemplateView):
    """
    Справочная информация о программе.

    Содержит:
    - Основные правила использования
    - Информацию о правах интеллектуальной собственности
    - Версию программы
    - Контактную информацию

    Требует аутентификации.
    """
    template_name = 'cards/about.html'
    login_url = 'admin:login'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        context['app_name'] = 'Система управления автотранспортом'
        context['version'] = '1.0.0'
        context['company'] = 'ЭХП РОСАТОМ, Автотранспортный цех 013'
        context['contact_email'] = 'support@example.com'
        context['copyright_year'] = '2025'

        return context


class FreeGarageNumbersView(LoginRequiredMixin, TemplateView):
    """
    Список свободных гаражных номеров.

    Показывает:
    - Используемые гаражные номера
    - Свободные гаражные номера
    - Возможность поиска по диапазону
    - Статус каждого номера

    Требует аутентификации.
    """
    template_name = 'cards/free_garage_numbers.html'
    login_url = 'admin:login'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        from .models import Vehicle

        # Получаем все используемые номера
        used_numbers = set(
            Vehicle.objects.filter(is_archived=False)
            .values_list('garage_number', flat=True)
            .distinct()
        )

        # Определяем диапазон (предположим 1-1000)
        all_numbers = set(range(1, 1001))
        free_numbers = sorted(list(all_numbers - used_numbers))

        context['used_numbers'] = sorted(list(used_numbers))
        context['free_numbers'] = free_numbers[:100]  # Первые 100 свободных для отображения
        context['total_free'] = len(free_numbers)
        context['total_used'] = len(used_numbers)
        context['total_possible'] = 1000

        return context


class MaintenanceBaseView(LoginRequiredMixin, TemplateView):
    """
    Главная страница раздела "Техническое обслуживание".

    Служит контейнером для подменю ТО:
    - Проведение ТО-2
    - ТС для проведения ТО-2
    - Графики (ГТО/ОСАГО, ГБО, ДПОГ)

    Требует аутентификации.
    """
    template_name = 'cards/maintenance_base.html'
    login_url = 'admin:login'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        from .models import MaintenanceSchedule

        # Статистика по ТО-2
        context['total_maintenance'] = MaintenanceSchedule.objects.count()
        context['scheduled_count'] = MaintenanceSchedule.objects.filter(status='scheduled').count()
        context['in_progress_count'] = MaintenanceSchedule.objects.filter(status='in_progress').count()
        context['overdue_count'] = MaintenanceSchedule.objects.filter(status='overdue').count()

        return context


class ReportsBaseView(LoginRequiredMixin, TemplateView):
    """
    Главная страница раздела "Отчёты".

    Служит контейнером для подменю отчётов:
    - Журнал аудита
    - Таблицы справочников
    - Экспорт данных

    Требует аутентификации.
    """
    template_name = 'cards/reports_base.html'
    login_url = 'admin:login'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        from .models import AuditLog, Vehicle, MaintenanceSchedule

        # Статистика
        context['total_vehicles'] = Vehicle.objects.filter(is_archived=False).count()
        context['archived_vehicles'] = Vehicle.objects.filter(is_archived=True).count()
        context['recent_audits'] = AuditLog.objects.all()[:10]
        context['total_audits'] = AuditLog.objects.count()
        context['maintenance_records'] = MaintenanceSchedule.objects.count()

        return context
